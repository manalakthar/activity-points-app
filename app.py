from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_from_directory, send_file
from datetime import datetime
import os
import requests as http_requests
from dotenv import load_dotenv
load_dotenv()
from database import (
    init_db, get_db, dict_cursor, get_student, get_student_by_email,
    get_submissions_by_student,
    sync_student_total_points,
    get_pending_submissions_for_mentor,
    get_pending_submissions_for_coordinator,
    get_pending_submissions_for_college,
    get_all_activities, get_mentor_for_student,
    get_all_assignments, get_current_calendar,
    get_all_calendars, get_eligible_students,
    advance_students, get_read_notification_keys, mark_notifications_read,
    get_assigned_students_with_activities
)

app = Flask(__name__)
app.secret_key = 'sjec_sap_secret_key'

# Folders
UPLOAD_FOLDER = 'uploads'
KNOWN_FACES_DIR = 'known_faces'
ALLOWED_PHOTO_EXTENSIONS = {'.jpg', '.jpeg', '.png'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(KNOWN_FACES_DIR, exist_ok=True)

# Supabase Storage config
SUPABASE_URL = os.environ.get('SUPABASE_URL', '').rstrip('/')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '')
FACE_BUCKET = 'face-photos'


def save_student_face_photo(student_id, face_photo):
    """Upload face photo to Supabase Storage and return the public URL."""
    if not face_photo or not face_photo.filename:
        return None, 'No photo selected.'
    extension = os.path.splitext(face_photo.filename)[1].lower()
    if extension not in ALLOWED_PHOTO_EXTENSIONS:
        return None, 'Please upload a JPG or PNG image.'

    file_bytes = face_photo.read()
    storage_path = f'{student_id}{extension}'
    content_type = 'image/jpeg' if extension in ('.jpg', '.jpeg') else 'image/png'

    if SUPABASE_URL and SUPABASE_KEY:
        # Upload to Supabase Storage (upsert so re-uploads work)
        upload_url = f'{SUPABASE_URL}/storage/v1/object/{FACE_BUCKET}/{storage_path}'
        headers = {
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'apikey': SUPABASE_KEY,
            'Content-Type': content_type,
            'x-upsert': 'true'
        }
        resp = http_requests.post(upload_url, headers=headers, data=file_bytes)
        print(f'[STORAGE] Upload response: {resp.status_code} {resp.text}')
        if resp.status_code in (200, 201):
            public_url = f'{SUPABASE_URL}/storage/v1/object/public/{FACE_BUCKET}/{storage_path}'
            # Also save locally for face recognition (fallback)
            local_path = os.path.join(KNOWN_FACES_DIR, storage_path)
            with open(local_path, 'wb') as f:
                f.write(file_bytes)
            return public_url, None
        else:
            print(f'[STORAGE] Upload failed: {resp.status_code} {resp.text}')
            # Fall back to local storage
    
    # Fallback: save locally only
    for ext in ALLOWED_PHOTO_EXTENSIONS:
        old_path = os.path.join(KNOWN_FACES_DIR, f'{student_id}{ext}')
        if os.path.exists(old_path):
            os.remove(old_path)
    path = os.path.join(KNOWN_FACES_DIR, storage_path)
    with open(path, 'wb') as f:
        f.write(file_bytes)
    return path, None


def resolve_student_mentor(student):
    mentor_id = get_mentor_for_student(student['student_id'], student['semester'])
    if not mentor_id:
        return None, None
    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('SELECT mentor_id, name FROM mentors WHERE mentor_id = %s', (mentor_id,))
    mentor = cur.fetchone()
    conn.close()
    if not mentor:
        return None, None
    return mentor['mentor_id'], mentor['name']


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


init_db()

# ============================================================
# HOME
# ============================================================

@app.route('/')
def home():
    return redirect(url_for('login'))

# ============================================================
# LOGIN
# ============================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')

        conn = get_db()
        cur = dict_cursor(conn)

        if role == 'student':
            cur.execute(
                'SELECT * FROM students WHERE email = %s AND password = %s',
                (email, password)
            )
            user = cur.fetchone()
            if user:
                session['user_id'] = user['student_id']
                session['role'] = 'student'
                session['name'] = user['name']
                conn.close()
                return redirect(url_for('student_dashboard'))

        elif role == 'mentor':
            cur.execute(
                'SELECT * FROM mentors WHERE email = %s AND password = %s',
                (email, password)
            )
            user = cur.fetchone()
            if user:
                session['user_id'] = user['mentor_id']
                session['role'] = 'mentor'
                session['name'] = user['name']
                conn.close()
                return redirect(url_for('mentor_dashboard'))

        elif role == 'coordinator':
            cur.execute(
                'SELECT * FROM coordinators WHERE email = %s AND password = %s',
                (email, password)
            )
            user = cur.fetchone()
            if user:
                session['user_id'] = user['coordinator_id']
                session['role'] = user['role']
                session['name'] = user['name']
                conn.close()
                if user['role'] == 'departmental':
                    return redirect(url_for('coordinator_dashboard'))
                else:
                    return redirect(url_for('college_dashboard'))

        conn.close()
        return render_template('login.html', error='Invalid email or password')

    return render_template('login.html')

# ============================================================
# LOGOUT
# ============================================================

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ============================================================
# FORGOT PASSWORD
# ============================================================

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        role = request.form.get('role')
        user_id = request.form.get('user_id')
        email = request.form.get('email')
        new_password = request.form.get('new_password')

        conn = get_db()
        cur = dict_cursor(conn)

        table = ''
        id_col = ''
        if role == 'student':
            table, id_col = 'students', 'student_id'
        elif role == 'mentor':
            table, id_col = 'mentors', 'mentor_id'
        elif role == 'coordinator':
            table, id_col = 'coordinators', 'coordinator_id'

        cur.execute(
            f'SELECT * FROM {table} WHERE {id_col} = %s AND email = %s',
            (user_id, email)
        )
        user = cur.fetchone()

        if user:
            cur.execute(
                f'UPDATE {table} SET password = %s WHERE {id_col} = %s',
                (new_password, user_id)
            )
            conn.commit()
            conn.close()
            return render_template('forgot_password.html',
                                   success='Password has been reset successfully!')

        conn.close()
        return render_template('forgot_password.html',
                               error='Invalid ID or Email combination')

    return render_template('forgot_password.html')

# ============================================================
# STUDENT REGISTRATION
# ============================================================

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        department = request.form.get('department')
        year = request.form.get('year')
        semester = request.form.get('semester')
        student_type = request.form.get('student_type')
        points_required = 75 if student_type == 'lateral' else 100

        face_photo = request.files.get('face_photo')
        face_photo_path = None

        if face_photo and face_photo.filename:
            face_photo_path, photo_error = save_student_face_photo(student_id, face_photo)
            if photo_error:
                return render_template('register.html', error=photo_error)

        try:
            conn = get_db()
            cur = dict_cursor(conn)
            cur.execute('''
                INSERT INTO students
                (student_id, name, email, password, department,
                 year, semester, student_type, points_required, face_photo_path)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (student_id, name, email, password, department,
                  year, semester, student_type, points_required, face_photo_path))
            conn.commit()
            conn.close()
            return redirect(url_for('login'))

        except Exception as e:
            print(f"Registration error: {e}")
            error_msg = 'Registration failed. A student with this USN or email already exists.'
            try:
                conn.close()
            except Exception:
                pass
            return render_template('register.html', error=error_msg)

    return render_template('register.html')

# ============================================================
# STUDENT DASHBOARD
# ============================================================

@app.route('/student/dashboard')
def student_dashboard():
    if session.get('role') != 'student':
        return redirect(url_for('login'))

    sync_student_total_points(session['user_id'])
    student = get_student(session['user_id'])
    submissions = get_submissions_by_student(session['user_id'])
    mentor_id, mentor_name = resolve_student_mentor(student)

    # Calculate points approved this academic year
    total_claimed_this_year = sum(
        sub['points_awarded']
        for sub in submissions
        if sub['year'] == student['year'] and sub['status'] == 'approved'
    )

    return render_template('student_dashboard.html',
                           student=student,
                           submissions=submissions,
                           mentor_assigned=mentor_id is not None,
                           mentor_name=mentor_name,
                           total_claimed_this_year=total_claimed_this_year)

# ============================================================
# STUDENT PROFILE
# ============================================================

@app.route('/student/profile', methods=['GET', 'POST'])
def student_profile():
    if session.get('role') != 'student':
        return redirect(url_for('login'))

    student_id = session['user_id']
    success_message = None
    error_message = None

    if request.method == 'POST':
        face_photo = request.files.get('face_photo')
        face_photo_path, photo_error = save_student_face_photo(student_id, face_photo)

        if photo_error:
            error_message = photo_error
        else:
            conn = get_db()
            cur = dict_cursor(conn)
            cur.execute(
                'UPDATE students SET face_photo_path = %s WHERE student_id = %s',
                (face_photo_path, student_id)
            )
            conn.commit()
            conn.close()
            success_message = 'Profile photo updated successfully.'

    sync_student_total_points(student_id)
    student = get_student(student_id)
    submissions = get_submissions_by_student(student_id)
    fp = student['face_photo_path']
    has_photo = bool(fp and (fp.startswith('http') or os.path.exists(fp)))

    return render_template('student_profile.html',
                           student=student,
                           submissions=submissions,
                           has_photo=has_photo,
                           success_message=success_message,
                           error_message=error_message)


@app.route('/student/profile/photo')
def student_profile_photo():
    if session.get('role') != 'student':
        return redirect(url_for('login'))

    student = get_student(session['user_id'])
    photo_path = student['face_photo_path'] if student else None
    if not photo_path:
        return '', 404

    # If it's a Supabase Storage URL, redirect directly to it
    if photo_path.startswith('http'):
        return redirect(photo_path)

    # Legacy: local file
    if not os.path.exists(photo_path):
        return '', 404
    directory = os.path.dirname(photo_path) or KNOWN_FACES_DIR
    filename = os.path.basename(photo_path)
    return send_from_directory(directory, filename)

# ============================================================
# ACTIVITIES PAGE
# ============================================================

@app.route('/activities')
def activities():
    all_activities = get_all_activities()
    return render_template('activities.html', activities=all_activities)

# ============================================================
# SUBMIT CLAIM
# ============================================================

@app.route('/student/submit', methods=['GET', 'POST'])
def submit_claim():
    if session.get('role') != 'student':
        return redirect(url_for('login'))

    student = get_student(session['user_id'])
    mentor_id, mentor_name = resolve_student_mentor(student)
    activities = get_all_activities()

    # Calculate current year approved points
    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('''
        SELECT COALESCE(SUM(points_awarded), 0) as total_claimed
        FROM submissions
        WHERE student_id = %s AND year = %s AND status = 'approved'
    ''', (session['user_id'], student['year']))
    total_claimed = cur.fetchone()['total_claimed']
    conn.close()
    reached_limit = (total_claimed >= 25)

    if request.method == 'POST':
        if not mentor_id:
            return render_template(
                'submit_claim.html',
                activities=activities,
                student=student,
                mentor_assigned=False,
                mentor_name=None,
                error=(
                    'You cannot submit this activity because no faculty '
                    'mentor has been assigned to you for this semester.'
                ),
            )

        activity_id = request.form.get('activity_id')
        role = request.form.get('role')
        organized_by = request.form.get('organized_by')
        activity_date = request.form.get('activity_date')
        duration_hours = request.form.get('duration_hours')
        points_claimed = request.form.get('points_claimed')
        protsaha_updated = 1 if request.form.get('protsaha_updated') else 0

        # Check annual approved limit
        if total_claimed >= 25:
            return render_template(
                'submit_claim.html',
                activities=activities,
                student=student,
                mentor_assigned=mentor_id is not None,
                mentor_name=mentor_name,
                total_claimed=total_claimed,
                reached_limit=reached_limit,
                error="You have already reached the limit of 25 approved points for this academic year. You cannot submit further claims."
            )

        # Fetch activity details to determine duplicate validation rules
        conn = get_db()
        cur = dict_cursor(conn)
        cur.execute('SELECT activity_name FROM activities WHERE activity_id = %s', (activity_id,))
        act = cur.fetchone()

        if not act:
            conn.close()
            return render_template(
                'submit_claim.html',
                activities=activities,
                student=student,
                mentor_assigned=mentor_id is not None,
                mentor_name=mentor_name,
                error='Invalid activity selected.'
            )

        is_class_rep = (act['activity_name'] == 'Class Representative')

        if is_class_rep:
            # Class Representative is limited to once per semester
            cur.execute('''
                SELECT COUNT(*) as count FROM submissions
                WHERE student_id = %s AND activity_id = %s AND semester = %s AND status != 'rejected'
            ''', (session['user_id'], activity_id, student['semester']))
            already_claimed = cur.fetchone()['count'] > 0
            if already_claimed:
                conn.close()
                return render_template(
                    'submit_claim.html',
                    activities=activities,
                    student=student,
                    mentor_assigned=mentor_id is not None,
                    mentor_name=mentor_name,
                    error='You have already claimed Class Representative points for this semester.'
                )
        else:
            # All other activities are limited to once per academic year
            cur.execute('''
                SELECT COUNT(*) as count FROM submissions
                WHERE student_id = %s AND activity_id = %s AND year = %s AND status != 'rejected'
            ''', (session['user_id'], activity_id, student['year']))
            already_claimed = cur.fetchone()['count'] > 0
            if already_claimed:
                conn.close()
                return render_template(
                    'submit_claim.html',
                    activities=activities,
                    student=student,
                    mentor_assigned=mentor_id is not None,
                    mentor_name=mentor_name,
                    error=f"You have already claimed points for '{act['activity_name']}' in this academic year (Year {student['year']}). It can only be claimed once a year."
                )
        conn.close()

        certificate = request.files.get('certificate')
        certificate_path = None
        extracted_text = None
        face_matched = 0

        if certificate:
            extension = os.path.splitext(certificate.filename)[1].lower()
            filename = f"{session['user_id']}_{datetime.now().strftime('%Y%m%d%H%M%S')}{extension}"
            certificate_path = os.path.join(UPLOAD_FOLDER, filename)
            certificate.save(certificate_path)

            try:
                from modules.ocr import extract_text
                extracted_text = extract_text(certificate_path) or None
            except Exception as e:
                print(f"OCR error: {e}")

            try:
                from modules.face_auth import verify_student
                result = verify_student(certificate_path, session['user_id'], KNOWN_FACES_DIR)
                face_matched = 1 if result else 0
            except Exception as e:
                print(f"Face recognition error: {e}")
                face_matched = 0

        conn = get_db()
        cur = dict_cursor(conn)
        cur.execute('''
            INSERT INTO submissions
            (student_id, activity_id, role, organized_by, activity_date,
             duration_hours, points_claimed, certificate_path, extracted_text,
             face_matched, protsaha_updated, status, mentor_id, submitted_date,
             semester, year)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s, %s, %s, %s)
        ''', (session['user_id'], activity_id, role, organized_by,
              activity_date, duration_hours, points_claimed,
              certificate_path, extracted_text, face_matched,
              protsaha_updated, mentor_id,
              datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              student['semester'], student['year']))
        conn.commit()
        conn.close()

        return redirect(url_for('student_dashboard'))

    return render_template(
        'submit_claim.html',
        activities=activities,
        student=student,
        mentor_assigned=mentor_id is not None,
        mentor_name=mentor_name,
        total_claimed=total_claimed,
        reached_limit=reached_limit,
    )

# ============================================================
# RESUBMIT CLAIM
# ============================================================

@app.route('/student/resubmit/<int:submission_id>', methods=['GET', 'POST'])
def resubmit_claim(submission_id):
    if session.get('role') != 'student':
        return redirect(url_for('login'))

    student = get_student(session['user_id'])
    mentor_id, mentor_name = resolve_student_mentor(student)
    activities = get_all_activities()

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('''
        SELECT s.*, a.activity_name FROM submissions s
        JOIN activities a ON s.activity_id = a.activity_id
        WHERE s.submission_id = %s AND s.student_id = %s
    ''', (submission_id, session['user_id']))
    submission = cur.fetchone()

    # Calculate approved points excluding this submission
    cur.execute('''
        SELECT COALESCE(SUM(points_awarded), 0) as total_claimed
        FROM submissions
        WHERE student_id = %s AND year = %s AND status = 'approved' AND submission_id != %s
    ''', (session['user_id'], student['year'], submission_id))
    total_claimed = cur.fetchone()['total_claimed']
    conn.close()

    reached_limit = (total_claimed >= 25)

    if not submission:
        return redirect(url_for('student_dashboard'))

    if request.method == 'POST':
        activity_id = request.form.get('activity_id')
        role = request.form.get('role')
        organized_by = request.form.get('organized_by')
        activity_date = request.form.get('activity_date')
        duration_hours = request.form.get('duration_hours')
        points_claimed = request.form.get('points_claimed')
        protsaha_updated = 1 if request.form.get('protsaha_updated') else 0

        # Check annual approved limit
        if total_claimed >= 25:
            return render_template(
                'submit_claim.html',
                activities=activities,
                student=student,
                mentor_assigned=mentor_id is not None,
                mentor_name=mentor_name,
                resubmit_submission=submission,
                total_claimed=total_claimed,
                reached_limit=reached_limit,
                error="You have already reached the limit of 25 approved points for this academic year. You cannot resubmit this claim."
            )
        # Fetch activity details to determine duplicate validation rules
        conn = get_db()
        cur = dict_cursor(conn)
        cur.execute('SELECT activity_name FROM activities WHERE activity_id = %s', (activity_id,))
        act = cur.fetchone()

        if not act:
            conn.close()
            return render_template(
                'submit_claim.html',
                activities=activities,
                student=student,
                mentor_assigned=mentor_id is not None,
                mentor_name=mentor_name,
                resubmit_submission=submission,
                error='Invalid activity selected.'
            )

        is_class_rep = (act['activity_name'] == 'Class Representative')

        if is_class_rep:
            # Class Representative is limited to once per semester
            cur.execute('''
                SELECT COUNT(*) as count FROM submissions
                WHERE student_id = %s AND activity_id = %s AND semester = %s 
                  AND status != 'rejected' AND submission_id != %s
            ''', (session['user_id'], activity_id, student['semester'], submission_id))
            already_claimed = cur.fetchone()['count'] > 0
            if already_claimed:
                conn.close()
                return render_template(
                    'submit_claim.html',
                    activities=activities,
                    student=student,
                    mentor_assigned=mentor_id is not None,
                    mentor_name=mentor_name,
                    resubmit_submission=submission,
                    error='You have already claimed Class Representative points for this semester.'
                )
        else:
            # All other activities are limited to once per academic year
            cur.execute('''
                SELECT COUNT(*) as count FROM submissions
                WHERE student_id = %s AND activity_id = %s AND year = %s 
                  AND status != 'rejected' AND submission_id != %s
            ''', (session['user_id'], activity_id, student['year'], submission_id))
            already_claimed = cur.fetchone()['count'] > 0
            if already_claimed:
                conn.close()
                return render_template(
                    'submit_claim.html',
                    activities=activities,
                    student=student,
                    mentor_assigned=mentor_id is not None,
                    mentor_name=mentor_name,
                    resubmit_submission=submission,
                    error=f"You have already claimed points for '{act['activity_name']}' in this academic year (Year {student['year']}). It can only be claimed once a year."
                )
        conn.close()

        certificate = request.files.get('certificate')
        certificate_path = submission['certificate_path']
        extracted_text = submission['extracted_text']
        face_matched = submission['face_matched'] or 0

        if certificate and certificate.filename:
            extension = os.path.splitext(certificate.filename)[1].lower()
            filename = f"{session['user_id']}_{datetime.now().strftime('%Y%m%d%H%M%S')}{extension}"
            certificate_path = os.path.join(UPLOAD_FOLDER, filename)
            certificate.save(certificate_path)

            try:
                from modules.ocr import extract_text
                extracted_text = extract_text(certificate_path) or None
            except Exception as e:
                print(f"OCR error: {e}")

            try:
                from modules.face_auth import verify_student
                result = verify_student(certificate_path, session['user_id'], KNOWN_FACES_DIR)
                face_matched = 1 if result else 0
            except Exception as e:
                print(f"Face recognition error: {e}")
                face_matched = 0

        conn = get_db()
        cur = dict_cursor(conn)
        cur.execute('''
            UPDATE submissions
            SET activity_id = %s, role = %s, organized_by = %s,
                activity_date = %s, duration_hours = %s, points_claimed = %s,
                certificate_path = %s, extracted_text = %s, face_matched = %s,
                protsaha_updated = %s, status = 'pending',
                rejection_note = NULL, submitted_date = %s,
                semester = %s, year = %s
            WHERE submission_id = %s AND student_id = %s
        ''', (activity_id, role, organized_by, activity_date, duration_hours,
              points_claimed, certificate_path, extracted_text, face_matched,
              protsaha_updated, datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              student['semester'], student['year'],
              submission_id, session['user_id']))
        conn.commit()
        conn.close()

        return redirect(url_for('student_dashboard'))

    return render_template(
        'submit_claim.html',
        activities=activities,
        student=student,
        mentor_assigned=mentor_id is not None,
        mentor_name=mentor_name,
        resubmit_submission=submission,
        total_claimed=total_claimed,
        reached_limit=reached_limit,
    )

# ============================================================
# MENTOR DASHBOARD
# ============================================================

@app.route('/mentor/dashboard')
def mentor_dashboard():
    if session.get('role') != 'mentor':
        return redirect(url_for('login'))

    submissions = get_pending_submissions_for_mentor(session['user_id'])
    assigned_students = get_assigned_students_with_activities(session['user_id'])
    return render_template('mentor_dashboard.html',
                           submissions=submissions,
                           assigned_students=assigned_students,
                           name=session['name'])

# ============================================================
# MENTOR REPORT VIEW (preview before download)
# ============================================================

@app.route('/mentor/view-report')
def mentor_view_report():
    if session.get('role') != 'mentor':
        return redirect(url_for('login'))

    submissions = get_pending_submissions_for_mentor(session['user_id'])
    assigned_students = get_assigned_students_with_activities(session['user_id'])

    # Compute summary totals
    total_approved = sum(len(s.get('approved_activities', [])) for s in assigned_students)
    total_pts_awarded = sum(
        act.get('points_awarded', 0) or 0
        for s in assigned_students
        for act in s.get('approved_activities', [])
    )

    generated_at = datetime.now().strftime('%d %B %Y, %I:%M %p')

    return render_template(
        'mentor_report_view.html',
        submissions=submissions,
        assigned_students=assigned_students,
        name=session['name'],
        generated_at=generated_at,
        total_approved=total_approved,
        total_pts_awarded=total_pts_awarded,
    )


# ============================================================
# MENTOR EXCEL EXPORT
# ============================================================

@app.route('/mentor/export-excel')
def mentor_export_excel():
    if session.get('role') != 'mentor':
        return redirect(url_for('login'))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    mentor_id   = session['user_id']
    mentor_name = session['name']
    submissions = get_pending_submissions_for_mentor(mentor_id)
    assigned_students = get_assigned_students_with_activities(mentor_id)

    wb = openpyxl.Workbook()

    # ── Helper styles ──────────────────────────────────────────
    header_font    = Font(bold=True, color='FFFFFF', size=11)
    header_fill    = PatternFill('solid', fgColor='1A56DB')   # blue
    sub_header_fill= PatternFill('solid', fgColor='2563EB')   # slightly lighter
    alt_fill       = PatternFill('solid', fgColor='EFF6FF')   # very light blue
    section_fill   = PatternFill('solid', fgColor='1E3A5F')   # dark navy for student rows
    section_font   = Font(bold=True, color='FFFFFF', size=10)
    center_align   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align     = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_side      = Side(style='thin', color='CBD5E0')
    thin_border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_header_row(ws, row_num, col_count, fill=None):
        fill = fill or header_fill
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font      = header_font
            cell.fill      = fill
            cell.alignment = center_align
            cell.border    = thin_border

    def style_data_cell(cell, row_num, alt=True):
        cell.alignment = left_align
        cell.border    = thin_border
        if alt and row_num % 2 == 0:
            cell.fill = alt_fill

    def auto_fit_columns(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ''
                    max_len = max(max_len, len(val))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # ══════════════════════════════════════════════════════════
    # SHEET 1 — Pending Submissions
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Pending Submissions'

    # Title row
    ws1.merge_cells('A1:G1')
    title_cell = ws1['A1']
    title_cell.value     = f'Pending Submissions — Mentor: {mentor_name}'
    title_cell.font      = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill      = PatternFill('solid', fgColor='1E3A5F')
    title_cell.alignment = center_align
    ws1.row_dimensions[1].height = 24

    # Generated date row
    ws1.merge_cells('A2:G2')
    date_cell = ws1['A2']
    date_cell.value     = f'Generated on: {datetime.now().strftime("%d %B %Y, %I:%M %p")}'
    date_cell.font      = Font(italic=True, size=9, color='64748B')
    date_cell.alignment = center_align
    ws1.row_dimensions[2].height = 16

    # Header row
    headers1 = ['Student Name', 'Student ID', 'Activity', 'Category', 'Role', 'Points Claimed', 'Submitted Date']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col, value=h)
    style_header_row(ws1, 3, len(headers1))
    ws1.row_dimensions[3].height = 18

    # Data rows
    if submissions:
        for row_num, sub in enumerate(submissions, start=4):
            values = [
                sub.get('student_name', ''),
                sub.get('student_id', ''),
                sub.get('activity_name', ''),
                sub.get('category', ''),
                str(sub.get('role', '')).capitalize(),
                sub.get('points_claimed', ''),
                str(sub.get('submitted_date', ''))[:10] if sub.get('submitted_date') else '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws1.cell(row=row_num, column=col, value=val)
                style_data_cell(cell, row_num)
    else:
        ws1.merge_cells('A4:G4')
        cell = ws1['A4']
        cell.value     = 'No pending submissions at this time.'
        cell.font      = Font(italic=True, color='718096')
        cell.alignment = center_align

    auto_fit_columns(ws1)

    # ══════════════════════════════════════════════════════════
    # SHEET 2 — Assigned Students & Activities
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Students & Activities')

    # Title
    ws2.merge_cells('A1:F1')
    t2 = ws2['A1']
    t2.value     = f'Assigned Students & Approved Activities — Mentor: {mentor_name}'
    t2.font      = Font(bold=True, size=13, color='FFFFFF')
    t2.fill      = PatternFill('solid', fgColor='1E3A5F')
    t2.alignment = center_align
    ws2.row_dimensions[1].height = 24

    ws2.merge_cells('A2:F2')
    d2 = ws2['A2']
    d2.value     = f'Generated on: {datetime.now().strftime("%d %B %Y, %I:%M %p")}'
    d2.font      = Font(italic=True, size=9, color='64748B')
    d2.alignment = center_align
    ws2.row_dimensions[2].height = 16

    col_headers2 = ['Student Name', 'Student ID', 'Year & Semester', 'Type', 'Points Earned', 'Points Required']
    for col, h in enumerate(col_headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
    style_header_row(ws2, 3, len(col_headers2))
    ws2.row_dimensions[3].height = 18

    current_row = 4
    act_headers = ['', 'Activity Name', 'Date', '', 'Points Awarded', '']

    for student in assigned_students:
        # Student summary row
        student_values = [
            student.get('name', ''),
            student.get('student_id', ''),
            f"Year {student.get('year', '')}, Sem {student.get('semester', '')}",
            str(student.get('student_type', '')).capitalize(),
            student.get('total_points', 0),
            student.get('points_required', 0),
        ]
        for col, val in enumerate(student_values, 1):
            cell = ws2.cell(row=current_row, column=col, value=val)
            cell.font      = section_font
            cell.fill      = section_fill
            cell.alignment = left_align
            cell.border    = thin_border
        ws2.row_dimensions[current_row].height = 18
        current_row += 1

        # Activity sub-header
        acts = student.get('approved_activities', [])
        if acts:
            sub_hdr_cols = {'B': 'Activity Name', 'C': 'Date', 'E': 'Points Awarded'}
            for col_letter, label in sub_hdr_cols.items():
                cell = ws2[f'{col_letter}{current_row}']
                cell.value     = label
                cell.font      = Font(bold=True, color='FFFFFF', size=10)
                cell.fill      = sub_header_fill
                cell.alignment = center_align
                cell.border    = thin_border
            ws2.row_dimensions[current_row].height = 16
            current_row += 1

            for act_row, act in enumerate(acts):
                act_date = str(act.get('submitted_date', ''))[:10] if act.get('submitted_date') else ''
                act_data = {'B': act.get('activity_name', ''), 'C': act_date, 'E': act.get('points_awarded', '')}
                for col_letter, val in act_data.items():
                    cell = ws2[f'{col_letter}{current_row}']
                    cell.value     = val
                    cell.border    = thin_border
                    cell.alignment = left_align
                    if act_row % 2 == 0:
                        cell.fill = alt_fill
                ws2.row_dimensions[current_row].height = 15
                current_row += 1
        else:
            cell = ws2[f'B{current_row}']
            cell.value     = 'No approved activities yet.'
            cell.font      = Font(italic=True, color='718096')
            cell.alignment = left_align
            current_row += 1

        current_row += 1  # blank spacer between students

    auto_fit_columns(ws2)

    # ── Save to in-memory buffer and send ─────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"mentor_{mentor_id}_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================================
# MENTOR REVIEW
# ============================================================

@app.route('/mentor/review/<int:submission_id>', methods=['GET', 'POST'])
def mentor_review(submission_id):
    if session.get('role') != 'mentor':
        return redirect(url_for('login'))

    conn = get_db()
    cur = dict_cursor(conn)

    if request.method == 'POST':
        action = request.form.get('action')
        rejection_note = request.form.get('rejection_note', '')
        status = 'mentor_approved' if action == 'approve' else 'rejected'

        cur.execute('''
            UPDATE submissions
            SET status = %s, reviewed_date = %s, rejection_note = %s
            WHERE submission_id = %s
        ''', (status,
              datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              rejection_note if action == 'reject' else None,
              submission_id))
        conn.commit()
        conn.close()
        return redirect(url_for('mentor_dashboard'))

    cur.execute('''
        SELECT s.*, a.activity_name, a.category, st.name AS student_name
        FROM submissions s
        JOIN activities a ON s.activity_id = a.activity_id
        JOIN students st ON s.student_id = st.student_id
        WHERE s.submission_id = %s
    ''', (submission_id,))
    submission = cur.fetchone()
    conn.close()

    return render_template('mentor_review.html', submission=submission)

# ============================================================
# COORDINATOR DASHBOARD
# ============================================================

@app.route('/coordinator/dashboard')
def coordinator_dashboard():
    if session.get('role') != 'departmental':
        return redirect(url_for('login'))

    submissions = get_pending_submissions_for_coordinator()
    return render_template('coordinator_dashboard.html',
                           submissions=submissions,
                           name=session['name'])

# ============================================================
# COORDINATOR REVIEW
# ============================================================

@app.route('/coordinator/review/<int:submission_id>', methods=['GET', 'POST'])
def coordinator_review(submission_id):
    if session.get('role') != 'departmental':
        return redirect(url_for('login'))

    conn = get_db()
    cur = dict_cursor(conn)

    if request.method == 'POST':
        action = request.form.get('action')
        points_awarded = int(request.form.get('points_awarded', 0))
        rejection_note = request.form.get('rejection_note', '')

        if action == 'approve':
            cur.execute('''
                UPDATE submissions
                SET status = 'approved', points_awarded = %s,
                    reviewed_date = %s, rejection_note = NULL
                WHERE submission_id = %s
            ''', (points_awarded,
                  datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  submission_id))

            cur.execute(
                'SELECT student_id FROM submissions WHERE submission_id = %s',
                (submission_id,)
            )
            submission = cur.fetchone()
            conn.commit()
            conn.close()
            sync_student_total_points(submission['student_id'])
        else:
            cur.execute('''
                UPDATE submissions
                SET status = 'rejected', reviewed_date = %s,
                    rejection_note = %s
                WHERE submission_id = %s
            ''', (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  rejection_note, submission_id))
            conn.commit()
            conn.close()

        return redirect(url_for('coordinator_dashboard'))

    cur.execute('''
        SELECT s.*, a.activity_name, a.category,
               a.max_points_participant, a.max_points_organizer,
               st.name AS student_name
        FROM submissions s
        JOIN activities a ON s.activity_id = a.activity_id
        JOIN students st ON s.student_id = st.student_id
        WHERE s.submission_id = %s
    ''', (submission_id,))
    submission = cur.fetchone()
    conn.close()

    return render_template('coordinator_review.html', submission=submission)

# ============================================================
# COLLEGE DASHBOARD
# ============================================================

@app.route('/college/dashboard')
def college_dashboard():
    if session.get('role') != 'college':
        return redirect(url_for('login'))

    submissions = get_pending_submissions_for_college()

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('SELECT * FROM students WHERE watch_list = 1')
    watchlist = cur.fetchall()
    conn.close()

    return render_template('college_dashboard.html',
                           submissions=submissions,
                           watchlist=watchlist,
                           name=session['name'])

# ============================================================
# COLLEGE FINAL APPROVAL
# ============================================================

@app.route('/college/review/<int:submission_id>', methods=['GET', 'POST'])
def college_review(submission_id):
    if session.get('role') != 'college':
        return redirect(url_for('login'))

    conn = get_db()
    cur = dict_cursor(conn)

    if request.method == 'POST':
        action = request.form.get('action')
        points_awarded = int(request.form.get('points_awarded', 0))

        if action == 'approve':
            cur.execute('''
                UPDATE submissions
                SET status = 'approved', points_awarded = %s,
                    reviewed_date = %s
                WHERE submission_id = %s
            ''', (points_awarded,
                  datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  submission_id))

            cur.execute(
                'SELECT student_id FROM submissions WHERE submission_id = %s',
                (submission_id,)
            )
            submission = cur.fetchone()
            conn.commit()
            conn.close()
            sync_student_total_points(submission['student_id'])
        else:
            cur.execute('''
                UPDATE submissions
                SET status = 'rejected', reviewed_date = %s
                WHERE submission_id = %s
            ''', (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), submission_id))
            conn.commit()
            conn.close()

        return redirect(url_for('college_dashboard'))

    cur.execute('''
        SELECT s.*, a.activity_name, a.category,
               a.max_points_participant, a.max_points_organizer,
               st.name AS student_name
        FROM submissions s
        JOIN activities a ON s.activity_id = a.activity_id
        JOIN students st ON s.student_id = st.student_id
        WHERE s.submission_id = %s
    ''', (submission_id,))
    submission = cur.fetchone()
    conn.close()

    return render_template('college_review.html', submission=submission)

# ============================================================
# WATCHLIST
# ============================================================

@app.route('/coordinator/watchlist')
def watchlist():
    if session.get('role') != 'departmental':
        return redirect(url_for('login'))

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('SELECT * FROM students WHERE watch_list = 1')
    students = cur.fetchall()
    conn.close()

    return render_template('watchlist.html', students=students)

# ============================================================
# ADMIN LOGIN
# ============================================================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        conn = get_db()
        cur = dict_cursor(conn)
        cur.execute(
            'SELECT * FROM admins WHERE email = %s AND password = %s',
            (email, password)
        )
        admin = cur.fetchone()
        conn.close()

        if admin:
            session['user_id'] = admin['admin_id']
            session['role'] = 'admin'
            session['name'] = admin['name']
            return redirect(url_for('admin_dashboard'))

        return render_template('admin_login.html', error='Invalid email or password')

    return render_template('admin_login.html')

# ============================================================
# ADMIN DASHBOARD
# ============================================================

@app.route('/admin/dashboard')
def admin_dashboard():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('SELECT * FROM students ORDER BY name')
    students = cur.fetchall()
    cur.execute('SELECT * FROM mentors ORDER BY name')
    mentors = cur.fetchall()
    cur.execute('SELECT * FROM coordinators ORDER BY name')
    coordinators = cur.fetchall()
    conn.close()

    return render_template('admin_dashboard.html',
                           students=students,
                           mentors=mentors,
                           coordinators=coordinators,
                           name=session['name'])

# ============================================================
# ADMIN ADD MENTOR
# ============================================================

@app.route('/admin/add_mentor', methods=['GET', 'POST'])
def admin_add_mentor():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        mentor_id = request.form.get('mentor_id')
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        department = request.form.get('department')

        try:
            conn = get_db()
            cur = dict_cursor(conn)
            cur.execute('''
                INSERT INTO mentors (mentor_id, name, email, password, department)
                VALUES (%s, %s, %s, %s, %s)
            ''', (mentor_id, name, email, password, department))
            conn.commit()
            conn.close()
            return redirect(url_for('admin_dashboard'))
        except Exception as e:
            return render_template('admin_add_mentor.html',
                                   error='Mentor ID or email already exists')

    return render_template('admin_add_mentor.html')

# ============================================================
# ADMIN ADD COORDINATOR
# ============================================================

@app.route('/admin/add_coordinator', methods=['GET', 'POST'])
def admin_add_coordinator():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        coordinator_id = request.form.get('coordinator_id')
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        department = request.form.get('department')

        try:
            conn = get_db()
            cur = dict_cursor(conn)
            cur.execute('''
                INSERT INTO coordinators
                (coordinator_id, name, email, password, role, department)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (coordinator_id, name, email, password, role, department))
            conn.commit()
            conn.close()
            return redirect(url_for('admin_dashboard'))
        except Exception as e:
            return render_template('admin_add_coordinator.html',
                                   error='Coordinator ID or email already exists')

    return render_template('admin_add_coordinator.html')

# ============================================================
# ADMIN DELETE USER
# ============================================================

@app.route('/admin/delete/<user_type>/<user_id>')
def admin_delete(user_type, user_id):
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()
    cur = dict_cursor(conn)

    if user_type == 'mentor':
        cur.execute('DELETE FROM mentors WHERE mentor_id = %s', (user_id,))
    elif user_type == 'coordinator':
        cur.execute('DELETE FROM coordinators WHERE coordinator_id = %s', (user_id,))
    elif user_type == 'student':
        cur.execute('DELETE FROM students WHERE student_id = %s', (user_id,))

    conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))

# ============================================================
# ADMIN — MENTOR ASSIGNMENTS
# ============================================================

@app.route('/admin/assignments')
def admin_assignments():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    assignments = get_all_assignments()

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('SELECT * FROM students ORDER BY name')
    students = cur.fetchall()
    cur.execute('SELECT * FROM mentors ORDER BY name')
    mentors = cur.fetchall()
    conn.close()

    return render_template('admin_assignments.html',
                           assignments=assignments,
                           students=students,
                           mentors=mentors)


@app.route('/admin/assign_mentor', methods=['POST'])
def assign_mentor():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    student_id = request.form.get('student_id')
    mentor_id = request.form.get('mentor_id')
    semester = request.form.get('semester')
    academic_year = request.form.get('academic_year')

    try:
        conn = get_db()
        cur = dict_cursor(conn)

        cur.execute('''
            SELECT assignment_id FROM mentor_assignments
            WHERE student_id = %s AND semester = %s
        ''', (student_id, semester))
        existing = cur.fetchone()

        if existing:
            cur.execute('''
                UPDATE mentor_assignments
                SET mentor_id = %s, academic_year = %s
                WHERE student_id = %s AND semester = %s
            ''', (mentor_id, academic_year, student_id, semester))
        else:
            cur.execute('''
                INSERT INTO mentor_assignments
                (student_id, mentor_id, semester, academic_year)
                VALUES (%s, %s, %s, %s)
            ''', (student_id, mentor_id, semester, academic_year))

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Assignment error: {e}")

    return redirect(url_for('admin_assignments'))


@app.route('/admin/delete_assignment/<int:assignment_id>')
def delete_assignment(assignment_id):
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('DELETE FROM mentor_assignments WHERE assignment_id = %s', (assignment_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_assignments'))

# ============================================================
# ADMIN — BULK MENTOR ASSIGNMENT
# ============================================================

@app.route('/admin/bulk_assign', methods=['GET', 'POST'])
def bulk_assign():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('SELECT * FROM mentors ORDER BY name')
    mentors = cur.fetchall()

    cur.execute('''
        SELECT DISTINCT department, semester
        FROM students
        ORDER BY department, semester
    ''')
    departments = cur.fetchall()

    selected_dept = request.args.get('department', '')
    selected_sem = request.args.get('semester', '')
    group_size = int(request.args.get('group_size', 30))

    students = []
    if selected_dept and selected_sem:
        cur.execute('''
            SELECT s.*,
                   ma.mentor_id AS assigned_mentor_id,
                   m.name AS assigned_mentor_name
            FROM students s
            LEFT JOIN mentor_assignments ma
                ON s.student_id = ma.student_id AND ma.semester = %s
            LEFT JOIN mentors m ON ma.mentor_id = m.mentor_id
            WHERE s.department = %s AND s.semester = %s
            ORDER BY s.student_id
        ''', (selected_sem, selected_dept, selected_sem))
        students = cur.fetchall()

    conn.close()

    groups = []
    for i in range(0, len(students), group_size):
        groups.append(list(students[i:i + group_size]))

    return render_template('admin_bulk_assign.html',
                           mentors=mentors,
                           departments=departments,
                           students=students,
                           groups=groups,
                           selected_dept=selected_dept,
                           selected_sem=selected_sem,
                           group_size=group_size)


@app.route('/admin/bulk_assign/save', methods=['POST'])
def bulk_assign_save():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    academic_year = request.form.get('academic_year')
    semester = request.form.get('semester')
    department = request.form.get('department')

    conn = get_db()
    cur = dict_cursor(conn)
    assigned_count = 0
    group_index = 0

    while True:
        mentor_id = request.form.get(f'group_{group_index}_mentor')
        student_ids = request.form.get(f'group_{group_index}_students')
        if mentor_id is None:
            break

        if student_ids and mentor_id:
            for student_id in student_ids.split(','):
                student_id = student_id.strip()
                if not student_id:
                    continue

                cur.execute('''
                    SELECT assignment_id FROM mentor_assignments
                    WHERE student_id = %s AND semester = %s
                ''', (student_id, semester))
                existing = cur.fetchone()

                if existing:
                    cur.execute('''
                        UPDATE mentor_assignments
                        SET mentor_id = %s, academic_year = %s
                        WHERE student_id = %s AND semester = %s
                    ''', (mentor_id, academic_year, student_id, semester))
                else:
                    cur.execute('''
                        INSERT INTO mentor_assignments
                        (student_id, mentor_id, semester, academic_year)
                        VALUES (%s, %s, %s, %s)
                    ''', (student_id, mentor_id, semester, academic_year))

                assigned_count += 1

        group_index += 1

    conn.commit()
    conn.close()
    return redirect(url_for('bulk_assign',
                            department=department,
                            semester=semester,
                            success=assigned_count))

# ============================================================
# ADMIN — ACADEMIC CALENDAR
# ============================================================

@app.route('/admin/calendar', methods=['GET', 'POST'])
def admin_calendar():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    if request.method == 'POST':
        semester = request.form.get('semester')
        academic_year = request.form.get('academic_year')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        is_current = 1 if request.form.get('is_current') else 0

        conn = get_db()
        cur = dict_cursor(conn)

        if is_current:
            cur.execute('UPDATE academic_calendar SET is_current = 0')

        cur.execute('''
            INSERT INTO academic_calendar
            (semester, academic_year, start_date, end_date, is_current, created_at)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (semester, academic_year, start_date, end_date, is_current,
              datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

        conn.commit()
        conn.close()
        return redirect(url_for('admin_calendar'))

    calendars = get_all_calendars()
    current = get_current_calendar()
    return render_template('admin_calendar.html', calendars=calendars, current=current)


@app.route('/admin/calendar/set_current/<int:calendar_id>')
def set_current_calendar(calendar_id):
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('UPDATE academic_calendar SET is_current = 0')
    cur.execute('UPDATE academic_calendar SET is_current = 1 WHERE calendar_id = %s', (calendar_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_calendar'))


@app.route('/admin/calendar/delete/<int:calendar_id>')
def delete_calendar(calendar_id):
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('DELETE FROM academic_calendar WHERE calendar_id = %s', (calendar_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_calendar'))

# ============================================================
# ADMIN — ADVANCE SEMESTER
# ============================================================

@app.route('/admin/advance_semester', methods=['GET', 'POST'])
def advance_semester():
    if session.get('role') != 'admin':
        return redirect(url_for('admin_login'))

    current_calendar = get_current_calendar()
    eligible_students = []
    success_message = None

    if current_calendar:
        eligible_students = get_eligible_students(current_calendar['semester'])

    if request.method == 'POST':
        selected_ids = request.form.getlist('student_ids')
        academic_year = request.form.get('academic_year')

        if selected_ids:
            advance_students(selected_ids, academic_year)
            success_message = (
                f"Successfully advanced {len(selected_ids)} students to next semester!"
            )
            if current_calendar:
                eligible_students = get_eligible_students(current_calendar['semester'])

    return render_template('admin_advance_semester.html',
                           current_calendar=current_calendar,
                           eligible_students=eligible_students,
                           success_message=success_message)

# ============================================================
# NOTIFICATIONS
# ============================================================

@app.route('/notifications')
def get_notifications():
    if not session.get('user_id'):
        return jsonify([])

    user_id = session['user_id']
    role = session.get('role')
    read_keys = get_read_notification_keys(user_id, role)

    conn = get_db()
    cur = dict_cursor(conn)
    notifications = []

    if role == 'student':
        cur.execute('''
            SELECT s.submission_id, a.activity_name, s.status, s.points_awarded
            FROM submissions s
            JOIN activities a ON s.activity_id = a.activity_id
            WHERE s.student_id = %s
            AND s.status IN ('approved', 'rejected', 'mentor_approved')
        ''', (user_id,))
        reviewed = cur.fetchall()

        for r in reviewed:
            notif_id = f"s-{r['submission_id']}-{r['status']}"
            if notif_id in read_keys:
                continue
            if r['status'] == 'approved':
                notifications.append({
                    'id': notif_id,
                    'message': f"✅ Your claim for '{r['activity_name']}' was approved! {r['points_awarded']} pts awarded.",
                    'type': 'success'
                })
            elif r['status'] == 'rejected':
                notifications.append({
                    'id': notif_id,
                    'message': f"❌ Your claim for '{r['activity_name']}' was rejected.",
                    'type': 'error'
                })
            elif r['status'] == 'mentor_approved':
                notifications.append({
                    'id': notif_id,
                    'message': f"🔄 Your claim for '{r['activity_name']}' is under coordinator review.",
                    'type': 'info'
                })

    elif role == 'mentor':
        cur.execute('''
            SELECT submission_id FROM submissions
            WHERE mentor_id = %s AND status = 'pending'
            ORDER BY submission_id
        ''', (user_id,))
        pending = cur.fetchall()

        unread_ids = [
            p['submission_id'] for p in pending
            if f"m-pending-{p['submission_id']}" not in read_keys
        ]
        if unread_ids:
            notif_id = 'm-pending-' + '-'.join(str(i) for i in unread_ids)
            notifications.append({
                'id': notif_id,
                'message': f"📋 You have {len(unread_ids)} pending submission(s) to review.",
                'type': 'warning'
            })

    elif role in ('departmental', 'college'):
        cur.execute('''
            SELECT submission_id FROM submissions
            WHERE status = 'mentor_approved'
            ORDER BY submission_id
        ''')
        pending = cur.fetchall()

        unread_ids = [
            p['submission_id'] for p in pending
            if f"c-pending-{p['submission_id']}" not in read_keys
        ]
        if unread_ids:
            notif_id = 'c-pending-' + '-'.join(str(i) for i in unread_ids)
            notifications.append({
                'id': notif_id,
                'message': f"📋 {len(unread_ids)} submission(s) are waiting for your review.",
                'type': 'warning'
            })

    conn.close()
    return jsonify(notifications)


@app.route('/notifications/read', methods=['POST'])
def read_notifications():
    if not session.get('user_id'):
        return jsonify({'ok': False}), 401

    data = request.get_json(silent=True) or {}
    keys = data.get('keys', [])
    expanded_keys = []

    for key in keys:
        if key.startswith('m-pending-'):
            suffix = key[len('m-pending-'):]
            if suffix:
                for submission_id in suffix.split('-'):
                    expanded_keys.append(f'm-pending-{submission_id}')
        elif key.startswith('c-pending-'):
            suffix = key[len('c-pending-'):]
            if suffix:
                for submission_id in suffix.split('-'):
                    expanded_keys.append(f'c-pending-{submission_id}')
        else:
            expanded_keys.append(key)

    mark_notifications_read(session['user_id'], session.get('role'), expanded_keys)
    return jsonify({'ok': True})


if __name__ == '__main__':
    app.run(debug=True)
